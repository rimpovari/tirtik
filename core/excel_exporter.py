"""
core/excel_exporter.py
-----------------------
Backtest sonuçlarını standart, formatlı Excel raporuna dönüştürür.
openpyxl gerektirir.
"""

from __future__ import annotations

import pandas as pd
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from core.order_simulator import Trade


def export_results(
    trades: list["Trade"],
    stats: dict,
    strategy_name: str,
    symbol: str,
    output_path: str | Path,
) -> None:
    """
    İki sayfalı Excel dosyası oluşturur:
      - 'Trades'    : Her işlemin detayı
      - 'Özet'      : Performans metrikleri + kümülatif R grafiği

    Parametre:
        trades        : simulate_trades() çıktısı
        stats         : compute_stats() çıktısı
        strategy_name : Rapor başlığında kullanılır
        symbol        : İşlem yapılan sembol
        output_path   : .xlsx çıktı yolu
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    trades_df = _trades_to_df(trades)
    summary_df = _stats_to_df(stats)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        trades_df.to_excel(writer, sheet_name="Trades", index=False)
        summary_df.to_excel(writer, sheet_name="Özet", index=False)

        _format_trades_sheet(writer, trades_df)
        _format_summary_sheet(writer, summary_df, trades_df, strategy_name, symbol)

    print(f"[Excel] Rapor kaydedildi: {output_path}")


# ── Yardımcı fonksiyonlar ─────────────────────────────────────────────────────

def _trades_to_df(trades: list) -> pd.DataFrame:
    rows = []
    for i, t in enumerate(trades, 1):
        risk = abs(t.entry_price - t.sl)
        rr = round(abs(t.tp - t.entry_price) / risk, 2) if risk else 0
        rows.append({
            "#": i,
            "Yön": t.direction.upper(),
            "Giriş Zamanı": t.entry_time,
            "Çıkış Zamanı": t.exit_time,
            "Giriş": round(t.entry_price, 5),
            "SL": round(t.sl, 5),
            "TP": round(t.tp, 5),
            "RR": rr,
            "Sonuç": t.result.upper(),
            "PnL (R)": round(t.pnl_r, 2),
            "Boyut": t.size,
        })
    return pd.DataFrame(rows)


def _stats_to_df(stats: dict) -> pd.DataFrame:
    label_map = {
        "total_trades": "Toplam İşlem",
        "wins": "Kazanan",
        "losses": "Kaybeden",
        "win_rate": "Kazanma Oranı",
        "total_r": "Toplam R",
        "profit_factor": "Profit Factor",
        "avg_win_r": "Ort. Kazanç (R)",
        "avg_loss_r": "Ort. Kayıp (R)",
        "max_drawdown_r": "Maks. Düşüş (R)",
        "expectancy_r": "Beklenti (R/işlem)",
    }
    rows = []
    for key, label in label_map.items():
        val = stats.get(key, "N/A")
        if key == "win_rate" and isinstance(val, float):
            val = f"{val:.1%}"
        rows.append({"Metrik": label, "Değer": val})
    return pd.DataFrame(rows)


def _format_trades_sheet(writer, df: pd.DataFrame) -> None:
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        ws = writer.sheets["Trades"]

        # Başlık satırı
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Win / Loss renklendirme
        win_fill = PatternFill("solid", fgColor="D1FAE5")
        loss_fill = PatternFill("solid", fgColor="FEE2E2")
        for row in ws.iter_rows(min_row=2):
            result_cell = next((c for c in row if c.column == 9), None)
            if result_cell:
                fill = win_fill if result_cell.value == "WIN" else (
                    loss_fill if result_cell.value == "LOSS" else None
                )
                if fill:
                    for cell in row:
                        cell.fill = fill

        # Kolon genişlikleri
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
    except ImportError:
        pass  # openpyxl yoksa formatlama atlanır


def _format_summary_sheet(writer, summary_df, trades_df, strategy_name, symbol) -> None:
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.chart import LineChart, Reference

        ws = writer.sheets["Özet"]

        # Başlık
        ws.insert_rows(1, 2)
        ws["A1"] = f"{strategy_name} — {symbol}"
        ws["A1"].font = Font(bold=True, size=14, color="1F2937")
        ws["A2"] = "Performans Özeti"
        ws["A2"].font = Font(italic=True, color="6B7280")

        # Metrik sütunları
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

        # Kümülatif R eğrisi verisi — Trades sayfasına yazdır
        ts = writer.sheets["Trades"]
        pnl_col = trades_df.columns.get_loc("PnL (R)") + 1
        n = len(trades_df)

        cum_col = trades_df.shape[1] + 1  # Kümülatif R sütunu
        ts.cell(1, cum_col, "Kümülatif R")
        cumulative = 0.0
        for r in range(2, n + 2):
            pnl_val = ts.cell(r, pnl_col).value or 0
            cumulative += pnl_val
            ts.cell(r, cum_col, round(cumulative, 2))

        # Grafik
        if n > 1:
            chart = LineChart()
            chart.title = "Kümülatif R Eğrisi"
            chart.style = 10
            chart.height = 12
            chart.width = 24
            data = Reference(ts, min_col=cum_col, min_row=1, max_row=n + 1)
            chart.add_data(data, titles_from_data=True)
            ws.add_chart(chart, "D4")
    except ImportError:
        pass
