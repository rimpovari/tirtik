"""
strategies/S02_Day_Candle_Catch/run_all.py
=======================================
30 parite x 2 LTF x 5 RR degeri backtest.

Cikti:
    results/{SEMBOL}.xlsx     -- her parite: Ozet sayfasi + her kombo icin trade sayfasi
    results/OZET_TUM.xlsx     -- matris + siralama + tum sonuclar
"""

from __future__ import annotations

import sys
import time
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT))

from strategies.S02_Day_Candle_Catch.backtest import run_single

SYMBOLS = [
    "EURUSD", "GBPUSD", "USDJPY", "USDCHF", "USDCAD", "AUDUSD", "NZDUSD",
    "XAUUSD", "XAGUSD",
    "EURGBP", "EURJPY", "EURCAD", "EURCHF", "EURAUD", "EURNZD",
    "GBPJPY", "GBPCHF", "GBPCAD", "GBPAUD", "GBPNZD",
    "AUDJPY", "AUDNZD", "AUDCAD", "AUDCHF",
    "CADJPY", "CADCHF", "CHFJPY", "NZDJPY", "NZDCAD", "NZDCHF",
]

LTFS = ["H1", "M15"]
RR_VALUES = [2.0, 3.0, 4.0, 5.0, 6.0]

RESULTS_DIR = Path(__file__).parent / "results"

def _combo_key(ltf: str, rr: float) -> str:
    return f"{ltf}_RR{rr:.0f}"

def _trades_to_df(trades) -> pd.DataFrame:
    rows = []
    for i, t in enumerate(trades, 1):
        risk = abs(t.entry_price - t.sl)
        rr   = round(abs(t.tp - t.entry_price) / risk, 2) if risk else 0
        rows.append({
            "#":           i,
            "Yon":         t.direction.upper(),
            "Giris Zaman": t.entry_time,
            "Cikis Zaman": t.exit_time,
            "Giris":       round(t.entry_price, 5),
            "SL":          round(t.sl, 5),
            "TP":          round(t.tp, 5),
            "RR":          rr,
            "Sonuc":       t.result.upper(),
            "PnL (R)":     round(t.pnl_r, 2),
            "Boyut":       t.size,
        })
    return pd.DataFrame(rows)

def _stats_row(symbol: str, ltf: str, rr: float, stats: dict) -> dict:
    return {
        "Sembol":       symbol,
        "LTF":          ltf,
        "Min_RR":       rr,
        "Combo":        _combo_key(ltf, rr),
        "Islem":        stats.get("total_trades", 0),
        "Kazanan":      stats.get("wins", 0),
        "Kaybeden":     stats.get("losses", 0),
        "WinRate":      round(stats.get("win_rate", 0), 4),
        "TotalR":       stats.get("total_r", 0),
        "ProfitFactor": stats.get("profit_factor", 0),
        "Beklenti_R":   stats.get("expectancy_r", 0),
        "MaxDD_R":      stats.get("max_drawdown_r", 0),
    }

def write_symbol_excel(symbol: str, combo_results: dict) -> None:
    # Hic sinyal yoksa Excel yazma
    if not any(trades for trades, _ in combo_results.values()):
        return

    path = RESULTS_DIR / f"{symbol}.xlsx"
    ozet_rows = []
    for key, (_, stats) in combo_results.items():
        if not stats:
            continue
        ltf, rr_str = key.split("_")[0], key.split("_RR")[1]
        ozet_rows.append({
            "Kombinasyon":  key,
            "LTF":          ltf,
            "Min_RR":       float(rr_str),
            "Islem":        stats.get("total_trades", 0),
            "Kazanan":      stats.get("wins", 0),
            "Kaybeden":     stats.get("losses", 0),
            "WinRate (%)":  round(stats.get("win_rate", 0) * 100, 1),
            "Toplam R":     stats.get("total_r", 0),
            "Profit Factor":stats.get("profit_factor", 0),
            "Beklenti R":   stats.get("expectancy_r", 0),
            "Max DD (R)":   stats.get("max_drawdown_r", 0),
        })

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        if ozet_rows:
            ozet_df = pd.DataFrame(ozet_rows).sort_values("Toplam R", ascending=False)
            ozet_df.to_excel(writer, sheet_name="Ozet", index=False)
            _format_ozet_sheet(writer.sheets["Ozet"])

        for key, (trades, stats) in combo_results.items():
            if not trades:
                continue
            sheet_name = key[:31]
            df = _trades_to_df(trades)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_trades_sheet(writer.sheets[sheet_name], df)

def _format_ozet_sheet(ws) -> None:
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        green_fill = PatternFill("solid", fgColor="DFFFD6")
        red_fill   = PatternFill("solid", fgColor="FFD6D6")
        total_r_col = [i for i, c in enumerate(ws[1], 1) if c.value == "Toplam R"]
        if total_r_col:
            col = total_r_col[0]
            for row in ws.iter_rows(min_row=2):
                val = row[col - 1].value
                if isinstance(val, (int, float)):
                    fill = green_fill if val > 0 else red_fill
                    for cell in row:
                        cell.fill = fill

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16
    except ImportError:
        pass

def _format_trades_sheet(ws, df: pd.DataFrame) -> None:
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        win_fill  = PatternFill("solid", fgColor="DFFFD6")
        loss_fill = PatternFill("solid", fgColor="FFD6D6")
        sonuc_col_idx = list(df.columns).index("Sonuc") + 1
        for row in ws.iter_rows(min_row=2):
            val = row[sonuc_col_idx - 1].value
            if val == "WIN":
                for cell in row:
                    cell.fill = win_fill
            elif val == "LOSS":
                for cell in row:
                    cell.fill = loss_fill

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16
    except ImportError:
        pass


def write_master_excel(all_rows: list[dict]) -> None:
    path = RESULTS_DIR / "OZET_TUM.xlsx"
    df   = pd.DataFrame(all_rows)

    if df.empty:
        print("[Master] Sonuc yok, OZET_TUM.xlsx olusturulamadi.")
        return

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_sorted = df.sort_values("TotalR", ascending=False)
        df_sorted.to_excel(writer, sheet_name="Tum_Sonuclar", index=False)
        _fmt_header(writer.sheets["Tum_Sonuclar"])
        
        _pivot_sheet(writer, df, "WinRate", "WinRate_Matrix", lambda v: f"{v:.1%}")
        _pivot_sheet(writer, df, "TotalR", "TotalR_Matrix")
        _pivot_sheet(writer, df, "ProfitFactor", "PF_Matrix")

        top = df_sorted.head(30).reset_index(drop=True)
        top.index += 1
        top.to_excel(writer, sheet_name="En_Iyi_30", index=True)
        _fmt_header(writer.sheets["En_Iyi_30"])

        best = (
            df.sort_values("TotalR", ascending=False)
              .groupby("Sembol")
              .first()
              .reset_index()
        )
        best.to_excel(writer, sheet_name="Sembol_En_Iyi", index=False)
        _fmt_header(writer.sheets["Sembol_En_Iyi"])

    print(f"\n[Master] OZET_TUM.xlsx kaydedildi: {path}")

def _pivot_sheet(writer, df: pd.DataFrame, value_col: str, sheet_name: str, fmt=None) -> None:
    try:
        pivot = df.pivot_table(index="Sembol", columns="Combo", values=value_col, aggfunc="first")
        if fmt:
            styled = pivot.applymap(fmt)
            styled.to_excel(writer, sheet_name=sheet_name)
        else:
            pivot.round(2).to_excel(writer, sheet_name=sheet_name)
        _fmt_header(writer.sheets[sheet_name])
        _color_pivot(writer.sheets[sheet_name])
    except Exception:
        pass

def _fmt_header(ws) -> None:
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        fill = PatternFill("solid", fgColor="1E3A5F")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 14
    except ImportError:
        pass

def _color_pivot(ws) -> None:
    try:
        from openpyxl.styles import PatternFill
        green = PatternFill("solid", fgColor="DFFFD6")
        red   = PatternFill("solid", fgColor="FFD6D6")
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                try:
                    val = float(str(cell.value).replace("%", ""))
                    cell.fill = green if val > 0 else red
                except (ValueError, TypeError):
                    pass
    except ImportError:
        pass

def main() -> None:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    total   = len(SYMBOLS) * len(LTFS) * len(RR_VALUES)
    done    = 0
    skipped = 0
    t_start = time.time()
    all_rows: list[dict] = []

    print(f"\n{'='*60}")
    print(f"  Day Candle Catch — Coklu Parametre Testi")
    print(f"  {len(SYMBOLS)} parite x {len(LTFS)} LTF x {len(RR_VALUES)} RR = {total} kombo")
    print(f"{'='*60}\n")

    for symbol in SYMBOLS:
        combo_results: dict[str, tuple] = {}
        for ltf in LTFS:
            for rr in RR_VALUES:
                done += 1
                key  = _combo_key(ltf, rr)
                pct  = done / total * 100

                try:
                    trades, stats = run_single(symbol, ltf, rr)
                except Exception:
                    trades, stats = [], {}
                    skipped += 1

                combo_results[key] = (trades, stats)

                if stats:
                    all_rows.append(_stats_row(symbol, ltf, rr, stats))
                    wr = stats.get("win_rate", 0)
                    tr = stats.get("total_r", 0)
                    n  = stats.get("total_trades", 0)
                    print(f"  [{done:>4}/{total}] {pct:4.0f}%  "
                          f"{symbol:<8} {ltf:<4} RR{rr:.0f}  "
                          f"{n:>3} islem  WR {wr:.0%}  R {tr:+.1f}")
                else:
                    print(f"  [{done:>4}/{total}] {pct:4.0f}%  "
                          f"{symbol:<8} {ltf:<4} RR{rr:.0f}  -- sinyal yok --")

        write_symbol_excel(symbol, combo_results)
        print(f"\n  -> {symbol}.xlsx kaydedildi\n")

    write_master_excel(all_rows)

    elapsed = time.time() - t_start
    print(f"\n{'='*60}")
    print(f"  Tamamlandi: {done} kombo ({skipped} atlandi)")
    print(f"  Sonuc uretilen: {len(all_rows)} kombinasyon")
    print(f"  Sure: {elapsed:.1f} sn")
    print(f"  Cikti: {RESULTS_DIR}")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
